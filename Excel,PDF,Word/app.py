import openpyxl
import os

workbook = openpyxl.load_workbook('./example.xlsx')
print(type(workbook))
sheet = workbook.get_sheet_by_name('Sheet1')
print(sheet)
cell = sheet['A1']
print(str(cell.value))
# print(sheet['A1'])
print(sheet['B1'].value)
print(sheet['C1'].value)
print(sheet.cell(row=1, column=1))

for i in range(1,8):
    print(i,sheet.cell(row=i, column =2).value)